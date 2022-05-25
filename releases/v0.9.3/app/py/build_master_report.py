#!/usr/bin/env python3
"""
BUILD_MASTER_REPORT_PY
Build a Master Report from all cases into a exercise.

=================================================================
This product is protected under U.S. Copyright Law.
Unauthorized reproduction is considered a criminal act.
(C) 2018-2021 VDI Technologies, LLC. All rights reserved. 
"""

__author__    = "Yoel Monsalve"
__date__      = "2021-11-12"
__modified__  = "2021-12-11"
__version__   = "0.9.1"
__copyright__ = "VDI Technologies, LLC"

import os          # os.path, os.stat, os.listdir
from sys import argv, stdin, stdout, stderr
import openpyxl    # read/write XLS
from openpyxl.formatting import Rule                            # conditional formatting
from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.styles.colors import Color
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

# custom modules
from helpers import getchar

def dump_data(fname, ws, row = 2):
    
    f = open(fname, 'r')
    if f:
        line = f.readline()
        while line:
            cols = line.strip().split('\t')
            for j in range(len(cols)):
                # numeric values
                if j >= column_index_from_string('C') and j <= column_index_from_string('K'):
                    try:
                        ws.cell(row=row, column=j+1).value = float(cols[j])
                    except:
                        # e.g. not a valid number
                        ws.cell(row=row, column=j+1).value = ''
                else:
                    ws.cell(row=row, column=j+1).value = cols[j]

            # increase row
            row += 1

            # feed another line from the CSV
            line = f.readline() 
        f.close()

    return row

# Conditional formatting with openpyxl
# 
# https://openpyxl.readthedocs.io/en/stable/formatting.html
def apply_conditional_format(ws, last_row = 2):

    if not range: return

    # clear any existing formatting
    ws.conditional_formatting = ConditionalFormattingList()

    badFill = PatternFill(start_color='CD5C5C',
                          end_color='CD5C5C',
                          fill_type='solid')
    
    dxf = DifferentialStyle(font=Font(italic=True, color="f0f0f0"), fill=badFill)
    
    # Rule to SPPR1 ( > 0.995 )
    rule_sppr1 = Rule(type='cellIs', dxf=dxf, operator='greaterThan', formula=['0.995'], stopIfTrue=False)
    # Rule to SPPR5 ( > 0.774 )
    rule_sppr5 = Rule(type='cellIs', dxf=dxf, operator='greaterThan', formula=['0.774'], stopIfTrue=False)
    # Rule to STD ( > 0.10 )
    rule_std = Rule(type='cellIs', dxf=dxf, operator='greaterThan', formula=['0.10'], stopIfTrue=False)

    try:
        ws.conditional_formatting.add('I2:I' + str(last_row), rule_sppr1)
        ws.conditional_formatting.add('J2:J' + str(last_row), rule_sppr5)
        ws.conditional_formatting.add('K2:K' + str(last_row), rule_std)
    except:
        # unable to apply conditional formatting
        pass

    # number format
    for row in ws['D2:K' + str(last_row)]:
        for cell in row:
            cell.alignment = Alignment(horizontal='right')

    for row in ws['D2:H' + str(last_row)]:
        for cell in row:
            cell.number_format = '0.00'

    for row in ws['I2:K' + str(last_row)]:
        for cell in row:
            cell.number_format = '0.0000'

    # header
    for row in ws['A1:M1']:
        for cell in row:
            cell.font = Font(bold=True)

def usage():
    prgname = os.path.basename(argv[0])
    stderr.write(f"USE:  {prgname} exercise_dir [xls_name]\n")

def run(exercise_dir, xls_name=''):
    """
    Runs the program properly

    @param  exercise_dir  str  The directory corresponding to the exercise to process
    @param  xls_name      str  The desired name to the output Excel file (defaults to 'Master-Report.xlsx')
    """

    # Excel file is supposed to be contained into the same directory
    # as the exercise dir
    xls_name = exercise_dir + '/' + xls_name

    # load workbook
    if not os.path.isfile(xls_name):
        wb = openpyxl.Workbook()
        wb.create_sheet('Unstables')
        wb.save(xls_name)
        wb.close()

    XLS = openpyxl.load_workbook(xls_name)
    if 'Unstables' in XLS:
        ws = XLS['Unstables']
    else:
        stderr.write(f"{argv[0]}: There is no sheet 'Unstables' in file '{xls_name}'")
        return 1
    row = 2    # first line to print on XLS file

    for d in os.listdir(exercise_dir):
        if os.path.isdir(exercise_dir + '/' + d):
            if os.path.isfile(exercise_dir + '/' + d + '/' + 'Failure-Report.csv'):
                fname = exercise_dir + '/' + d + '/' + 'Failure-Report.csv'
                row = dump_data(fname, ws, row)

    apply_conditional_format(ws, row-1)
    print(f"\nSaving file {xls_name}")
    XLS.save(xls_name)
    XLS.close()

def main():

    if len(argv) < 2:
        usage()
        return 1
    elif len(argv) < 3:
        exercise_dir = argv[1]
        xls_name     = 'Master-Report.xlsx'
    else:
        exercise_dir = argv[1]
        xls_name     = argv[2]
    
    #exercise_dir = "../../../data/sample_data"
    #xls_name = './Master-Report.xlsx'

    # run the program
    run(exercise_dir, xls_name)

    # success
    return 0

if __name__ == "__main__":
    main()