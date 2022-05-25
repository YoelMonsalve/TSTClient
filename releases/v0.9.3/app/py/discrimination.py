#!/usr/bin/env python3
"""
 * DISCRIMINATION_PY
 * Classify the different event failures by types, e.g.
 * sppr1, sppr2, std, tc.
 *
 * =================================================================
 * This product is protected under U.S. Copyright Law.
 * Unauthorized reproduction is considered a criminal act.
 * (C) 2018-2021 VDI Technologies, LLC. All rights reserved. 
"""

__author__    = "Yoel Monsalve"
__date__      = "2021-12-15"
__modified__  = "2021-12-19"
__version__   = "0.9.1"
__copyright__ = "VDI Technoslogies, LLC"


import os
import sys
from sys import stdin, stdout, stderr, argv
import subprocess
from datetime import datetime, timedelta
from time import sleep
import signal
import re         # regex
from shutil import copy, move, rmtree        # file operations

# pie chart
from matplotlib import pyplot as plt

# OpenPyXl
import openpyxl    # read/write XLS
from openpyxl import load_workbook
#from openpyxl.formatting import Rule                  # conditional formatting
#from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.styles import NamedStyle, Font, PatternFill, Border, Side, Alignment
from openpyxl.styles.colors import Color
#from openpyxl.styles.differential import DifferentialStyle
#from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

# custom modules
from helpers import getchar, is_win, is_posix
from helpers import list_folders, list_files

def comb(n: int, elems: list) -> list:
    """Makes all ordered combinations from n elems into the
    array elems
    """
    results = []
    if n < 1: return
    for j in range(0, len(elems)-n+1):
        part_comb = [elems[j]]     # a partial combination

        if not elems[j+1:] is None and n>1:
            for r in comb(n-1, elems[j+1:]):
                final_comb = part_comb + r
                results.append(final_comb)
        else:
            results.append(part_comb)
    
    return results

def make_verdict(file_path):
    """Makes a verdict over the violations (if any) in the event referred by
    the file path.
    """
    verdict = {
        'type'  : '',
        'sppr1' : '',
        'sppr5' : '',
        'std'   : '',
        '1.20'  : '',
        '0.70'  : '',
        'stable': ''
    }

    # regular patterns
    pattern_angl = re.compile("\\s*ANGL[0-9]*\\[.*\\][0-9]*");
    pattern_volt = re.compile("\\s*VOLT\\s*[0-9]*\\s\\[.*\\][0-9]*");

    if not os.path.isfile(file_path):
        stderr.write(f"classify_violations: Not a regular file: '{file_path}'\n")
        return {}

    f = open(file_path, 'r')
    for line in f:
        tokens = line.strip().split('\t')

        if len(tokens) >= 12:
            filename     = tokens[0]
            channel_name = tokens[1] 

            # channels for ANGL
            if pattern_angl.match(channel_name):
                """ ANGL
                =================================================
                tokens[0]  : Item #
                tokens[1]  : channel_name
                tokens[2]  : tc
                tokens[3]  : Min peak
                tokens[4]  : Max peak
                tokens[5]  : peak 1
                tokens[6]  : peak 2
                tokens[7]  : peak 6
                tokens[8]  : SPPR1
                tokens[9]  : SPPR5
                tokens[10] : std
                tokens[11] : stable?
                """
                verdict['type'] = 'ANGL'

                # test: SPPR1
                if tokens[8] and tokens[8] != '-':
                    sppr1 = float(tokens[8])
                    if (sppr1 > 0.95):         # failed if SPPR1 > 0.95
                        verdict['sppr1']  = 'FAILED'
                        verdict['stable'] = 'N'
                # test: SPPR5
                if tokens[9] and tokens[9] != '-':
                    sppr5 = float(tokens[9])
                    if (sppr1 > 0.774):         # failed if SPPR5 > 0.774
                        verdict['sppr5']  = 'FAILED'
                        verdict['stable'] = 'N'
                # test: ANGLE DEVIATION
                if tokens[10] and tokens[10] != '-':
                    std = float(tokens[10])
                    if (std > 0.10):         # failed if STD > 0.10
                        verdict['std']  = 'FAILED'
                        verdict['stable'] = 'N'

            # channels for VOLT
            if pattern_volt.match(channel_name):
                """ VOLT
                =================================================
                tokens[0]  : Item #
                tokens[1]  : channel_name
                tokens[2]  : tc
                tokens[3]  : 1.20 pu
                tokens[4]  : 0.70 pu
                tokens[11] : stable?
                """
                verdict['type'] = 'VOLT'

                # test: 1.20 PU
                if tokens[3] and tokens[3] != '-':
                    test_120 = str(tokens[3])
                    if (test_120 != 'OK'):
                        verdict['1.20']  = 'FAILED'
                        verdict['stable'] = 'N'
                # test: 0.70 PU
                if tokens[4] and tokens[4] != '-':
                    test_070 = str(tokens[4])
                    if (test_070 != 'OK'):
                        verdict['0.70']  = 'FAILED'
                        verdict['stable'] = 'N'

    f.close()

    if verdict['stable'] != 'N': verdict['stable'] = 'Y'
    return verdict

def scan_exercise(exercise_dir: str) -> list:
    """
    Looks through a parti1cular exercise, and calls the function to classify
    the violations for each event into such exercise.
    Returns a list of verdicts, one to each event located.

    @param  exercise_dir   str  The directory for the exercise (e.g. /path/to/work/20xx_TPL)
    """

    results = []
    if os.path.isdir(f"{exercise_dir}"):
        cases = list_folders(f"{exercise_dir}")
        for cs in cases:
            events = list_files(f"{exercise_dir}/{cs}/output/summary")
            cnt = 0
            unstables = 0
            for e in events:
                path = f"{exercise_dir}/{cs}/output/summary/{e}"
                verdict = make_verdict(path)
                results.append(verdict)

                if verdict:
                    verdict['case'] = cs
                    # converts P_X_YY_ZZZ_out.summary into P_X-YY_ZZZ
                    if e[-8:] == '.summary':
                        verdict['event'] = e.split('.summary')[0]
                    else:
                        verdict['event'] = e

                # --- debug ---
                cnt +=1
                if verdict['stable'] == 'N':
                    unstables+=1

    return results

def make_classifiers() -> list:
    """Makes all the classifiers, by combination of ANGL and VOLT
    criteria"""

    classifiers_A = ['sppr1', 'sppr5', 'std']     # ANGL
    classifiers_B = ['1.20', '0.70']              # VOLT
    classifiers   = []                            # all classifiers

    combs_A = []
    for j in range(1, len(classifiers_A)+1):      # j = 1, 2, .. len(classifiers_A)
        r = comb(j, classifiers_A)
        combs_A += r

    combs_B = []
    for j in range(1, len(classifiers_B)+1):      # j = 1, 2, .. len(classifiers_B)
        r = comb(j, classifiers_B)
        combs_B += r

    classifiers = combs_A + combs_B

    # now crossing {combs_A(angl)} X {combs_B(volt)}
    for a in combs_A:
        for b in combs_B:
            classifiers.append(a + b)

    classifiers.append(['stable'])
    return classifiers

def filter_results(results: list, classifier: list) -> list:
    """
    Returns a list of events that match FAILURE by the given classifier

    @param   results      list   A list of verdicts (as returned by scan_exercise())
    @param   classifier   list   A list of classifiers to be applied to pick up an event
                                 (e.g. ['sppr1','std','1.20'])
    """
    if not results or not list: return []

    filtered = []
    for r in results:
        PICK = True
        for tag in classifier:
            if tag != 'stable':
                if tag in r and r[tag] != 'FAILED':
                    PICK = False
                    break
            else:
                PICK = (r['stable'] == 'Y')
        if PICK:
            filtered.append(r)

    return filtered

def makeXLSReport(exercise_dir: str, unstables: list, dest_dir: str = ''):
    """
    [Description]

    @param   exercise_dir   str   [param1_desc]
    @param   unstables      list   [param2_desc]
    @param   dest_dir       str    [param2_desc]
    """

    if not exercise_dir or not unstables: return
    if not dest_dir: dest_dir = exercise_dir

    # Now, filling up the Excel report
    #source_dir = os.path.realpath(os.path.dirname(os.path.realpath(__file__)) + '/../templates/')
    source_dir = os.path.realpath(__file__ + '/../../templates/')

    book     = 'Judgement(template).xlsx'
    book2    = 'Judgement.xlsx'

    XLS_name  = source_dir + '/' + book      # the template book
    XLS_name2 = dest_dir + '/' + book2       # the final book

    wb = openpyxl.load_workbook(XLS_name)
    if not 'Summary' in wb:
        stderr.write(f"{os.path.basename(argv[0])}: Not a sheet named 'Summary' in {XLS_name}")
    if not 'Model' in wb:
        stderr.write(f"{os.path.basename(argv[0])}: Not a sheet named 'Summary' in {XLS_name}")
    for ws in wb.worksheets:
        if ws.title != 'Summary' and ws.title != 'Model':
            wb.remove(ws)

    ws = wb['Summary']
    exercise_name = os.path.basename(exercise_dir)
    ws['B1'] = exercise_name

    ws['A1'].font = Font(
        size = 16,
        bold = True,
        color='808080')
    ws['A1'].style = 'Headline 1'
    ws['B1'].style = 'Headline 1'

    bd1 = Side(border_style='double', color="ff808080")

    ws['A3'].font = Font(name='Cambria', bold=True)
    ws['A3'].border = Border(bottom=bd1)
    ws['B3'].font = Font(name='Cambria', bold=True)
    ws['B3'].border = Border(bottom=bd1)
    ws['C3'].font = Font(name='Cambria', bold=True)
    ws['C3'].border = Border(bottom=bd1)

    # fill data in data table
    Z = make_classifiers()
    i = 0
    for z in Z:
        #print(f"{i+1}: {z}")
        i+=1

    MasterTable = []
    for z in Z:
        filtered = filter_results(unstables, z)
        MasterTable.append({
            'classifier': z,
            'classifier2': ' & '.join(z),     # a string, e.g. ['sppr1', 'sppr5', '1.20'] -> 'sppr1 & sppr5 & 1.20'
            'events': filtered,
            'quantity': len(filtered)
        })

    # sort by quantity of each occurrence, in decreasing order
    #summary = [[t['classifier2'], t['quantity']] for t in MasterTable]
    MasterTable.sort(key = lambda t: t['quantity'], reverse=True)

    n_uns = len(unstables)
    i = 0
    for t in MasterTable:
        ws['A' + str(i+4)] = t['classifier2']
        ws['B' + str(i+4)] = t['quantity']
        ws['C' + str(i+4)] = round(t['quantity']/n_uns*100, 1)

        ws['B' + str(i+4)].alignment = Alignment(horizontal='right')
        ws['C' + str(i+4)].alignment = Alignment(horizontal='right')
        #ws['C' + str(i+4)].style = 'Percent'
        i += 1

    # filling data for individual classifiers
    for t in MasterTable:
        ws_model = wb['Model']
        ws = wb.copy_worksheet(ws_model)
        ws.title = t['classifier2']
        ws['A1'] = t['classifier2']
        ws['A1'].style = 'Headline 1'
        ws['B1'].style = 'Calculation'
        ws['A2'].font = Font(bold=True)
        ws['B2'].font = Font(bold=True)

        row = 3
        for e in t['events']:
            ws['A' + str(row)] = e['case']
            ws['B' + str(row)] = e['event']
            row += 1

    # hiding pattern sheet 'Model'
    wb['Model'].sheet_state = 'hidden'

    print(f"Saving file {XLS_name2}")
    wb.save(XLS_name2)

def pieChart(MasterTable: list, n_events: int, exercise_dir: str):

    # Making a data taking the first items that add more than 90%
    data = [[t['classifier2'], t['quantity']] for t in MasterTable]
    total = 0
    n_data = sum(d[1] for d in data)
    slices = []
    labels = []
    i = 0
    while total <= 0.95*n_data and i < len(data):
        total += data[i][1]

        labels.append(data[i][0].upper())
        slices.append(data[i][1])
        i += 1
    labels.append('OTHERS')
    slices.append(n_data - total)
    explode = [0] * len(labels)
    if 'STABLE' in labels:
        i = labels.index('STABLE')
        explode[i] = 0.10

    plt.style.use('fivethirtyeight')
    #ax.pie(slices, explode=explode, labels=labels, autopct='%1.1f%%', pctdistance=0.75, 
    #       wedgeprops={'edgecolor': 'black'}, startangle=90, 
    #       textprops={'size': 'medium', 'color': 'black'})
    fig, ax = plt.subplots(dpi=120)
    ax.pie(slices, explode=explode, labels=labels, autopct='%1.1f%%', startangle=0)
    #ax.legend(title = 'Cases discrimination')
    ax.axis('equal')

    plt.title(f"Engineering Judgment analysis results for {n_events} events", 
              color='black', loc='center', pad=20, fontsize=12)
    plt.tight_layout()
    #plt.show()

    exercise_name = os.path.basename(exercise_dir)
    figname = f"{exercise_dir}/eng_results_{exercise_name}.png"
    print("Saving figure to " + figname)
    fig.savefig(figname, transparent=False, 
                dpi=600, bbox_inches="tight")

def usage():
    prgname = os.path.basename(argv[0])
    stderr.write(f"USE:  {prgname} exercise_dir [xls_name] [dest_dir]\n")

def main():
    """
    [Description]

    @param   exercise_dir   str   [param1_desc]
    @param   xls_name       str   [param2_desc]
    """

    if len(argv) < 2:
        usage()
        return 1
    elif len(argv) < 3:
        exercise_dir = argv[1]
        xls_name     = 'Judgement.xlsx'
        dest_dir     = exercise_dir
    elif len(argv) < 4:
        exercise_dir = argv[1]
        xls_name     = argv[2]
        dest_dir     = exercise_dir
    else:
        exercise_dir = argv[1]
        xls_name     = argv[2]
        dest_dir     = argv[3]
    
    #exercise_dir  = '/home/tst/work/2019_TPL'
    unstables = scan_exercise(exercise_dir)
    
    Z = make_classifiers()
    i = 0
    for z in Z:
        #print(f"{i+1}: {z}")
        i+=1

    MasterTable = []
    for z in Z:
        filtered = filter_results(unstables, z)
        MasterTable.append({
            'classifier': z,
            'classifier2': ' & '.join(z),     # a string, e.g. ['sppr1', 'sppr5', '1.20'] -> 'sppr1 & sppr5 & 1.20'
            'events': filtered,
            'quantity': len(filtered)
        })
    MasterTable.sort(key = lambda x: x['quantity'], reverse = True)
    summary = [[t['classifier2'], t['quantity']] for t in MasterTable]
    #summary.sort(key = lambda x: x[1], reverse=True)
    n_uns = len(unstables)
    
    """print(f"Total: {n_uns} unstables.\nDiscrimination:\n================================================================")
    for a in summary:
        print(f"{a[0]:40s}    {a[1]:4d}  {a[1]/n_uns*100:.1f}%")
    """

    # build the report in Excel format
    # ========================================
    makeXLSReport(exercise_dir, unstables, dest_dir)

    # buid the Pie Chart
    # ========================================
    pieChart(MasterTable, len(unstables), dest_dir)

if __name__ == "__main__":
    main()