#!/usr/bin/env python

"""
This script will read the XLSX file and will export the sheet "OUTDATA"
as a separate CSV file

SYNTAX:
========

  prg xlsx_infile csv_outfile

Required Packages:
==================

  openpyxl
  
__________________________________
 This is part of the TST Project.
 This product is protected under U.S. Copyright Law.
 Unauthorized reproduction is considered a criminal act.
(C) 2018-2919 VDI Technologies, LLC. All rights reserved
"""

__author__    = "Yoel Monsalve"
__date__      = "January, 2019"
__modified__  = "June, 2021"
__version__   = ""
__copyright__ = "VDI Technologies, LLC"


__author__ = "Yoel Monsalve"
__date__   = "November, 2019"
__mail__   = "yymonsalve@gmail.com"
__copyright__ = "DVI Technologies, LLC"

import sys, os
from sys import argv, stdout, stderr, exit
from datetime import datetime

import openpyxl

# C-like ferror() function !!!
def ferror( msg ):

	stderr.write( "{:s}: {:s}\n".format( argv[0], msg ) )

def usage():

	stderr.write( "USAGE:  {:s} input_folder\n\n".format( argv[0] ) + \
	  "Folder that contains all the xlsx files to convert.\n" )
 
"""
Converts the given (xlsx) file to a CSV file with the
same name 
"""
def convert_file( filename ):
	
	if not "." in filename:
		ferror( "not extension in filename: '{:s}'".format( filename ) )
		return
		
	k = filename.find( "." )
	fname = filename[:k]
	ext = filename[k+1:]
	
	if not (ext == 'xls' or ext == 'xlsx' or ext == 'xlsm' ):
		#ferror( "not allowed extension: '{:s}'".format( filename ) )
		return
	
	t1 = datetime.now()
	print( "Converting '{:s}' ...".format( filename ) )
	
	WB = openpyxl.load_workbook( fname + '.' + ext )
	S = None
	# looks for a sheet named "OUTDATA"
	for sheet in WB:
		if sheet.title.upper() == "OUTDATA":
			S = sheet
			break
	if not S: 
		# if not found
		WB.close()
		return
	
	csv = open( fname + '.csv', 'w' )
	
	# auto-detecting range in the sheet
	MAX_ROW = 0
	MAX_COL = 0
	i = 1
	j = 1
	
	for row in S.iter_rows( ):
		line = ''
		# writing row as a line in the CSV
		for v in row:
			if line:
				line += ','
			line += str( v.value )
		csv.write( line + '\n' )
	"""
	# === deprecated ===
	while S.cell(i,1).value:
		line = ''
		# writing row as a line in the CSV
		if MAX_COL == 0:
			j = 1
			while S.cell(i,j).value:
				#print( "({:d},{:d}):".format(i,j), S.cell(i,j).value )
				if line:
					line += ','
				line += str( S.cell(i,j).value )
				MAX_COL = j
				j += 1
		else:
			for j in range(1, MAX_COL + 1 ):
				#print( "({:d},{:d}):".format(i,j), S.cell(i,j).value )
				if line:
					line += ','
				line += str( S.cell(i,j).value )
						
		csv.write( line + '\n' )
		i += 1
	"""
		
	csv.close()
	WB.close()
	t2 = datetime.now()
	print( "    saved to ==> '{:s}'".format( fname + '.csv' ) )
	print( "    elapsed: " + str( t2 - t1 ) )
	
def main():

	if len( argv ) < 2:
		usage()
		exit( 1 )
		
	input_dir = argv[1]
	for file in os.listdir( input_dir ):
		if os.path.isfile( input_dir + '/' + file ):
			convert_file( input_dir + '/' + file )

	return 0		# success

if __name__ == "__main__":

	main()
	

